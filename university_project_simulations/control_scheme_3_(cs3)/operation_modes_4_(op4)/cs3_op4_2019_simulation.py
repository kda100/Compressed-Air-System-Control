# -*- coding: utf-8 -*-
"""
Created on Tue Feb 18 16:19:09 2020
@author: kda_1

***************************************************************************************************************************************
This program loads an excel workbook from the same directory the script is contained in, called 'compressed_air_data_12_bar.xlsx'
which contains the 12 bar compressed air data from a leading car manufacturer's 12 bar compressed air system. The program then loads
the sheet 'air_volume_stacked_2019' and iterates over the column that contains the 12 bar compressed air data for 2019. For each
air demand value contained in the sheet the program produces and stores air output and power output values for each compressor 
based on a compressor control sequencing logic defined by 'setpoints' and 'active_compressors' in the program. The program then writes the 
power output and air output values into two separate sheets called 'compressors_power_output' and 'compressors_air_output' in an 
excel workbook called 'cs3_op4_2019_results.xlsx' located in the same directory as the script.

***************************************************************************************************************************************
2019
Maximum Flow of System - FAD (m3/hr) = 8672
Maximum Air Demand of Plant - FAD (m3/hr) = 5340

Control Sequence:

Setpoints (m3/hr):
s1 - 947
s2 - 1894
s3 - 1902
s4 - 2841
s5 - 2849
s6 - 3796
s7 - 4743
s8 - 6069
s9 - 8673

Setpoint Boundary (m3/hr)    Active Compressors
0-s1                         [C5]
s1-s2                        [C1, C5]
s2-s3                        [C1, C2, C5]
s3-s4                        [C6, C5]
s4-s5                        [C1, C2, C3, C5]
s5-s6                        [C1, C6, C5]
s6-s7                        [C1, C2, C6, C5]
s7-s8                        [C1, C2, C3, C6, C5]
s8-s9                        [C1, C2, C3, C4, C6, C5]

C1 = LoadUnload: max_power = 125, max_flow = 947
C2 = LoadUnload: max_power = 125, max_flow = 947
C3 = LoadUnload: max_power = 125, max_flow = 947
C4 = LoadUnload: max_power = 177, max_flow = 1326
C5 = InletModulation: max_power = 384, max_flow = 2604
C6 = LoadUnload: max_power = 274, max_flow = 1902

***************************************************************************************************************************************
"""
import openpyxl
import cas_modules.compressor_models1 as cm1

C1 = cm1.LoadUnload(125, 947) #create LoadUnload compressor object for C1
C2 = cm1.LoadUnload(125, 947) #create LoadUnload compressor object for C2
C3 = cm1.LoadUnload(125, 947) #create LoadUnload compressor object for C3
C4 = cm1.LoadUnload(177, 1326) #create LoadUnload compressor object for C4
C5 = cm1.InletModulation(384, 2604) #create InletModulation compressor object for C5
C6 = cm1.LoadUnload(274, 1902) #create LoadUnload compressor object for C6

C1_power_output = [] #list object to store power output for compressor C1
C1_air_output = [] #list object to store air output for compressor C1
C2_power_output = [] #list object to store power output for compressor C2
C2_air_output = [] #list object to store air output for compressor C2
C3_power_output = [] #list object to store power output for compressor C3
C3_air_output = [] #list object to store air output for compressor C3
C4_power_output = [] #list object to store power output for compressor C4
C4_air_output = [] #list object to store air output for compressor C4
C5_power_output = [] #list object to store power output for compressor C5
C5_air_output = [] #list object to store air output for compressor C5
C6_power_output = [] #list object to store power output for compressor C6
C6_air_output = [] #list object to store air output for compressor C6

Compressed_Air_Data_12_Bar = openpyxl.load_workbook('compressed_air_data_12_bar.xlsx') #open compressed air system data workbook in current directory
Air_Volume_Stacked_2019 = Compressed_Air_Data_12_Bar['air_volume_stacked_2019'] #calls sheet 'air_volume_stacked_2019' in workbook
for row in range(1,8089):
    air_demand = Air_Volume_Stacked_2019.cell(row = row, column = 7).value #iterates over all rows in column
    if air_demand == 0: #boundary
        active_compressors=[] #active_compressors for compressor.power_output and compressor.air_output
        control_input = False #control input for compressor.power_output and compressor.air_output
        C1_power_output.append(C1.power_output(control_input, air_demand, active_compressors)) #appends C1 power output reading to C1_power_output list
        C1_air_output.append(C1.air_output(control_input, air_demand, active_compressors)) #appends C1 air output reading to C1_air_output list
        C2_power_output.append(C2.power_output(control_input, air_demand, active_compressors)) #appends C2 power output reading to C2_power_output list
        C2_air_output.append(C2.air_output(control_input, air_demand, active_compressors)) #appends C2 air output reading to C2_air_output list
        C3_power_output.append(C3.power_output(control_input, air_demand, active_compressors)) #appends C3 power output reading to C3_power_output list
        C3_air_output.append(C3.air_output(control_input, air_demand, active_compressors)) #appends C3 air output reading to C3_air_output list
        C4_power_output.append(C4.power_output(control_input, air_demand, active_compressors)) #appends C4 power output reading to C4_power_output list
        C4_air_output.append(C4.air_output(control_input, air_demand, active_compressors)) #appends C4 air output reading to C4_air_output list
        C5_power_output.append(C5.power_output(control_input, air_demand, active_compressors)) #appends C5 power output reading to C5_power_output list
        C5_air_output.append(C5.air_output(control_input, air_demand, active_compressors)) #appends C5 air output reading to C5_air_output list
        C6_power_output.append(C6.power_output(control_input, air_demand, active_compressors)) #appends C6 power output reading to C6_power_output list
        C6_air_output.append(C6.air_output(control_input, air_demand, active_compressors)) #appends C6 air output reading to C6_air_output list
        
    compressors = [C1, C2, C3, C4, C5, C6] #compressor objects list
    s1 = compressors[0].maximum_flow #setpoint
        
    if 0<air_demand<=s1: #boundary
        active_compressors = [compressors[4]] #active_compressors for compressor.power_output and compressor.air_output
        control_input = True #control input for compressor.power_output and compressor.air_output
        C5_power_output.append(C5.power_output(control_input, air_demand, active_compressors)) #appends C5 power output reading to C5_power_output list
        C5_air_output.append(C5.air_output(control_input, air_demand, active_compressors)) #appends C5 air output reading to C5_air_output list
        control_input = False #control input for compressor.power_output and compressor.air_output
        C1_power_output.append(C1.power_output(control_input, air_demand, active_compressors)) #appends C1 power output reading to C1_power_output list
        C1_air_output.append(C1.air_output(control_input, air_demand, active_compressors)) #appends C1 air output reading to C1_air_output list
        C2_power_output.append(C2.power_output(control_input, air_demand, active_compressors)) #appends C2 power output reading to C2_power_output list
        C2_air_output.append(C2.air_output(control_input, air_demand, active_compressors)) #appends C2 air output reading to C2_air_output list
        C3_power_output.append(C3.power_output(control_input, air_demand, active_compressors)) #appends C3 power output reading to C3_power_output list
        C3_air_output.append(C3.air_output(control_input, air_demand, active_compressors)) #appends C3 air output reading to C3_air_output list
        C4_power_output.append(C4.power_output(control_input, air_demand, active_compressors)) #appends C4 power output reading to C4_power_output list
        C4_air_output.append(C4.air_output(control_input, air_demand, active_compressors)) #appends C4 air output reading to C4_air_output list
        C6_power_output.append(C6.power_output(control_input, air_demand, active_compressors)) #appends C6 power output reading to C6_power_output list
        C6_air_output.append(C6.air_output(control_input, air_demand, active_compressors)) #appends C6 air output reading to C6_air_output list

    compressors = [C1, C2, C3, C4, C5, C6] #compressor objects list
    s1 = compressors[0].maximum_flow #setpoint
    s2 = compressors[0].maximum_flow + compressors[1].maximum_flow #setpoint
    
    if s1<air_demand<=s2: #boundary
        active_compressors = [compressors[0], compressors[4]] #active_compressors for compressor.power_output and compressor.air_output
        control_input = True #control input for compressor.power_output and compressor.air_output
        C1_power_output.append(C1.power_output(control_input, air_demand, active_compressors)) #appends C1 power output reading to C1_power_output list
        C1_air_output.append(C1.air_output(control_input, air_demand, active_compressors)) #appends C1 air output reading to C1_air_output list
        C5_power_output.append(C5.power_output(control_input, air_demand, active_compressors)) #appends C5 power output reading to C5_power_output list
        C5_air_output.append(C5.air_output(control_input, air_demand, active_compressors)) #appends C5 air output reading to C5_air_output list
        control_input = False #control input for compressor.power_output and compressor.air_output
        C2_power_output.append(C2.power_output(control_input, air_demand, active_compressors)) #appends C2 power output reading to C2_power_output list
        C2_air_output.append(C2.air_output(control_input, air_demand, active_compressors)) #appends C2 air output reading to C2_air_output list
        C3_power_output.append(C3.power_output(control_input, air_demand, active_compressors)) #appends C3 power output reading to C3_power_output list
        C3_air_output.append(C3.air_output(control_input, air_demand, active_compressors)) #appends C3 air output reading to C3_air_output list
        C4_power_output.append(C4.power_output(control_input, air_demand, active_compressors)) #appends C4 power output reading to C4_power_output list
        C4_air_output.append(C4.air_output(control_input, air_demand, active_compressors)) #appends C4 air output reading to C4_air_output list
        C6_power_output.append(C6.power_output(control_input, air_demand, active_compressors)) #appends C6 power output reading to C6_power_output list
        C6_air_output.append(C6.air_output(control_input, air_demand, active_compressors)) #appends C6 air output reading to C6_air_output list
        
    compressors = [C1, C2, C3, C4, C5, C6] #compressor objects list
    s2 = compressors[0].maximum_flow + compressors[1].maximum_flow #setpoint
    s3 = compressors[5].maximum_flow #setpoint
    
    if s2<air_demand<=s3: #boundary
        active_compressors = [compressors[0], compressors[1], compressors[4]] #active_compressors for compressor.power_output and compressor.air_output
        control_input = True #control input for compressor.power_output and compressor.air_output
        C1_power_output.append(C1.power_output(control_input, air_demand, active_compressors)) #appends C1 power output reading to C1_power_output list
        C1_air_output.append(C1.air_output(control_input, air_demand, active_compressors)) #appends C1 air output reading to C1_air_output list
        C2_power_output.append(C2.power_output(control_input, air_demand, active_compressors)) #appends C2 power output reading to C2_power_output list
        C2_air_output.append(C2.air_output(control_input, air_demand, active_compressors)) #appends C2 air output reading to C2_air_output list
        C5_power_output.append(C5.power_output(control_input, air_demand, active_compressors)) #appends C5 power output reading to C5_power_output list
        C5_air_output.append(C5.air_output(control_input, air_demand, active_compressors)) #appends C5 air output reading to C5_air_output list
        control_input = False #control input for compressor.power_output and compressor.air_output
        C3_power_output.append(C3.power_output(control_input, air_demand, active_compressors)) #appends C3 power output reading to C3_power_output list
        C3_air_output.append(C3.air_output(control_input, air_demand, active_compressors)) #appends C3 air output reading to C3_air_output list
        C4_power_output.append(C4.power_output(control_input, air_demand, active_compressors)) #appends C4 power output reading to C4_power_output list
        C4_air_output.append(C4.air_output(control_input, air_demand, active_compressors)) #appends C4 air output reading to C4_air_output list
        C6_power_output.append(C6.power_output(control_input, air_demand, active_compressors)) #appends C6 power output reading to C6_power_output list
        C6_air_output.append(C6.air_output(control_input, air_demand, active_compressors)) #appends C6 air output reading to C6_air_output list
        
    compressors = [C1, C2, C3, C4, C5, C6] #compressor objects list
    s3 = compressors[5].maximum_flow #setpoint
    s4 = compressors[0].maximum_flow + compressors[1].maximum_flow + compressors[2].maximum_flow #setpoint
        
    if s3<air_demand<=s4: #boundary
        active_compressors = [compressors[5], compressors[4]] #active_compressors for compressor.power_output and compressor.air_output
        control_input = True #control input for compressor.power_output and compressor.air_output
        C6_power_output.append(C6.power_output(control_input, air_demand, active_compressors)) #appends C6 power output reading to C6_power_output list
        C6_air_output.append(C6.air_output(control_input, air_demand, active_compressors)) #appends C6 air output reading to C6_air_output list
        C5_power_output.append(C5.power_output(control_input, air_demand, active_compressors)) #appends C5 power output reading to C5_power_output list
        C5_air_output.append(C5.air_output(control_input, air_demand, active_compressors)) #appends C5 air output reading to C5_air_output list
        control_input = False #control input for compressor.power_output and compressor.air_output
        C1_power_output.append(C1.power_output(control_input, air_demand, active_compressors)) #appends C1 power output reading to C1_power_output list
        C1_air_output.append(C1.air_output(control_input, air_demand, active_compressors)) #appends C1 air output reading to C1_air_output list
        C2_power_output.append(C2.power_output(control_input, air_demand, active_compressors)) #appends C2 power output reading to C2_power_output list
        C2_air_output.append(C2.air_output(control_input, air_demand, active_compressors)) #appends C2 air output reading to C2_air_output list
        C3_power_output.append(C3.power_output(control_input, air_demand, active_compressors)) #appends C3 power output reading to C3_power_output list
        C3_air_output.append(C3.air_output(control_input, air_demand, active_compressors)) #appends C3 air output reading to C3_air_output list
        C4_power_output.append(C4.power_output(control_input, air_demand, active_compressors)) #appends C4 power output reading to C4_power_output list
        C4_air_output.append(C4.air_output(control_input, air_demand, active_compressors)) #appends C4 air output reading to C4_air_output list
        
    compressors = [C1, C2, C3, C4, C5, C6] #compressor objects list
    s4 = compressors[0].maximum_flow + compressors[1].maximum_flow + compressors[2].maximum_flow #setpoint
    s5 = compressors[5].maximum_flow + compressors[0].maximum_flow #setpoint
        
    if s4<air_demand<=s5: #boundary
        active_compressors = [compressors[0], compressors[1], compressors[2], compressors[4]] #active_compressors for compressor.power_output and compressor.air_output
        control_input = True #control input for compressor.power_output and compressor.air_output
        C1_power_output.append(C1.power_output(control_input, air_demand, active_compressors)) #appends C1 power output reading to C1_power_output list
        C1_air_output.append(C1.air_output(control_input, air_demand, active_compressors)) #appends C1 air output reading to C1_air_output list
        C2_power_output.append(C2.power_output(control_input, air_demand, active_compressors)) #appends C2 power output reading to C2_power_output list
        C2_air_output.append(C2.air_output(control_input, air_demand, active_compressors)) #appends C2 air output reading to C2_air_output list
        C3_power_output.append(C3.power_output(control_input, air_demand, active_compressors)) #appends C3 power output reading to C3_power_output list
        C3_air_output.append(C3.air_output(control_input, air_demand, active_compressors)) #appends C3 air output reading to C3_air_output list
        C5_power_output.append(C5.power_output(control_input, air_demand, active_compressors)) #appends C5 power output reading to C5_power_output list
        C5_air_output.append(C5.air_output(control_input, air_demand, active_compressors)) #appends C5 air output reading to C5_air_output list
        control_input = False #control input for compressor.power_output and compressor.air_output
        C4_power_output.append(C4.power_output(control_input, air_demand, active_compressors)) #appends C4 power output reading to C4_power_output list
        C4_air_output.append(C4.air_output(control_input, air_demand, active_compressors)) #appends C4 air output reading to C4_air_output list
        C6_power_output.append(C6.power_output(control_input, air_demand, active_compressors)) #appends C6 power output reading to C6_power_output list
        C6_air_output.append(C6.air_output(control_input, air_demand, active_compressors)) #appends C6 air output reading to C6_air_output list
        
    compressors = [C1, C2, C3, C4, C5, C6] #compressor objects list
    s5 = compressors[5].maximum_flow + compressors[0].maximum_flow #setpoint
    s6 = compressors[0].maximum_flow + compressors[1].maximum_flow + compressors[5].maximum_flow #setpoint
        
    if s5<air_demand<=s6: #boundary
        active_compressors = [compressors[0], compressors[5], compressors[4]] #active_compressors for compressor.power_output and compressor.air_output
        control_input = True #control input for compressor.power_output and compressor.air_output
        C1_power_output.append(C1.power_output(control_input, air_demand, active_compressors)) #appends C1 power output reading to C1_power_output list
        C1_air_output.append(C1.air_output(control_input, air_demand, active_compressors)) #appends C1 air output reading to C1_air_output list
        C5_power_output.append(C5.power_output(control_input, air_demand, active_compressors)) #appends C5 power output reading to C5_power_output list
        C5_air_output.append(C5.air_output(control_input, air_demand, active_compressors)) #appends C5 air output reading to C5_air_output list
        C6_power_output.append(C6.power_output(control_input, air_demand, active_compressors)) #appends C6 power output reading to C6_power_output list
        C6_air_output.append(C6.air_output(control_input, air_demand, active_compressors)) #appends C6 air output reading to C6_air_output list
        control_input = False #control input for compressor.power_output and compressor.air_output
        C2_power_output.append(C2.power_output(control_input, air_demand, active_compressors)) #appends C2 power output reading to C2_power_output list
        C2_air_output.append(C2.air_output(control_input, air_demand, active_compressors)) #appends C2 air output reading to C2_air_output list
        C3_power_output.append(C3.power_output(control_input, air_demand, active_compressors)) #appends C3 power output reading to C3_power_output list
        C3_air_output.append(C3.air_output(control_input, air_demand, active_compressors)) #appends C3 air output reading to C3_air_output list)
        C4_power_output.append(C4.power_output(control_input, air_demand, active_compressors)) #appends C4 power output reading to C4_power_output list
        C4_air_output.append(C4.air_output(control_input, air_demand, active_compressors)) #appends C4 air output reading to C4_air_output list
        
    compressors = [C1, C2, C3, C4, C5, C6] #compressor objects list
    s6 = compressors[0].maximum_flow + compressors[1].maximum_flow + compressors[5].maximum_flow #setpoint
    s7 = compressors[0].maximum_flow + compressors[1].maximum_flow + compressors[2].maximum_flow + compressors[5].maximum_flow #setpoint
    
    if s6<air_demand<=s7: #boundary
        active_compressors = [compressors[0], compressors[1], compressors[5], compressors[4]] #active_compressors for compressor.power_output and compressor.air_output
        control_input = True #control input for compressor.power_output and compressor.air_output
        C1_power_output.append(C1.power_output(control_input, air_demand, active_compressors)) #appends C1 power output reading to C1_power_output list
        C1_air_output.append(C1.air_output(control_input, air_demand, active_compressors)) #appends C1 air output reading to C1_air_output list
        C2_power_output.append(C2.power_output(control_input, air_demand, active_compressors)) #appends C2 power output reading to C2_power_output list
        C2_air_output.append(C2.air_output(control_input, air_demand, active_compressors)) #appends C2 air output reading to C2_air_output list
        C5_power_output.append(C5.power_output(control_input, air_demand, active_compressors)) #appends C5 power output reading to C5_power_output list
        C5_air_output.append(C5.air_output(control_input, air_demand, active_compressors)) #appends C5 air output reading to C5_air_output list
        C6_power_output.append(C6.power_output(control_input, air_demand, active_compressors)) #appends C6 power output reading to C6_power_output list
        C6_air_output.append(C6.air_output(control_input, air_demand, active_compressors)) #appends C6 air output reading to C6_air_output list
        control_input = False
        C3_power_output.append(C3.power_output(control_input, air_demand, active_compressors)) #appends C3 power output reading to C3_power_output list
        C3_air_output.append(C3.air_output(control_input, air_demand, active_compressors)) #appends C3 air output reading to C3_air_output list)
        C4_power_output.append(C4.power_output(control_input, air_demand, active_compressors)) #appends C4 power output reading to C4_power_output list
        C4_air_output.append(C4.air_output(control_input, air_demand, active_compressors)) #appends C4 air output reading to C4_air_output list
        
    compressors = [C1, C2, C3, C4, C5, C6] #compressor objects list
    s7 = compressors[0].maximum_flow + compressors[1].maximum_flow + compressors[2].maximum_flow + compressors[5].maximum_flow #setpoint
    s8 = compressors[0].maximum_flow + compressors[1].maximum_flow + compressors[2].maximum_flow + compressors[5].maximum_flow + compressors[3].maximum_flow #setpoint
    
    if s7<air_demand<=s8: #boundary
        active_compressors = [compressors[0], compressors[1], compressors[2], compressors[5], compressors[4]] #active_compressors for compressor.power_output and compressor.air_output
        control_input = True #control input for compressor.power_output and compressor.air_output
        C1_power_output.append(C1.power_output(control_input, air_demand, active_compressors)) #appends C1 power output reading to C1_power_output list
        C1_air_output.append(C1.air_output(control_input, air_demand, active_compressors)) #appends C1 air output reading to C1_air_output list
        C2_power_output.append(C2.power_output(control_input, air_demand, active_compressors)) #appends C2 power output reading to C2_power_output list
        C2_air_output.append(C2.air_output(control_input, air_demand, active_compressors)) #appends C2 air output reading to C2_air_output list
        C3_power_output.append(C3.power_output(control_input, air_demand, active_compressors)) #appends C3 power output reading to C3_power_output list
        C3_air_output.append(C3.air_output(control_input, air_demand, active_compressors)) #appends C3 air output reading to C3_air_output list
        C5_power_output.append(C5.power_output(control_input, air_demand, active_compressors)) #appends C5 power output reading to C5_power_output list
        C5_air_output.append(C5.air_output(control_input, air_demand, active_compressors)) #appends C5 air output reading to C5_air_output list
        C6_power_output.append(C6.power_output(control_input, air_demand, active_compressors)) #appends C6 power output reading to C6_power_output list
        C6_air_output.append(C6.air_output(control_input, air_demand, active_compressors)) #appends C6 air output reading to C6_air_output list
        control_input = False #control input for compressor.power_output and compressor.air_output
        C4_power_output.append(C4.power_output(control_input, air_demand, active_compressors)) #appends C4 power output reading to C4_power_output list
        C4_air_output.append(C4.air_output(control_input, air_demand, active_compressors)) #appends C4 air output reading to C4_air_output list
        
    compressors = [C1, C2, C3, C4, C5, C6] #compressor objects list
    s8 = compressors[0].maximum_flow + compressors[1].maximum_flow + compressors[2].maximum_flow + compressors[5].maximum_flow + compressors[3].maximum_flow #setpoint
    s9 = compressors[0].maximum_flow + compressors[1].maximum_flow + compressors[2].maximum_flow + compressors[5].maximum_flow + compressors[3].maximum_flow + compressors[4].maximum_flow #setpoint
    
    if s8<air_demand<=s9: #boundary
        active_compressors = [compressors[0], compressors[1], compressors[2], compressors[5], compressors[3], compressors[4]] #active_compressors for compressor.power_output and compressor.air_output
        control_input = True #control input for compressor.power_output and compressor.air_output
        C1_power_output.append(C1.power_output(control_input, air_demand, active_compressors)) #appends C1 power output reading to C1_power_output list
        C1_air_output.append(C1.air_output(control_input, air_demand, active_compressors)) #appends C1 air output reading to C1_air_output list
        C2_power_output.append(C2.power_output(control_input, air_demand, active_compressors)) #appends C2 power output reading to C2_power_output list
        C2_air_output.append(C2.air_output(control_input, air_demand, active_compressors)) #appends C2 air output reading to C2_air_output list
        C3_power_output.append(C3.power_output(control_input, air_demand, active_compressors)) #appends C3 power output reading to C3_power_output list
        C3_air_output.append(C3.air_output(control_input, air_demand, active_compressors)) #appends C3 air output reading to C3_air_output list
        C4_power_output.append(C4.power_output(control_input, air_demand, active_compressors)) #appends C4 power output reading to C4_power_output list
        C4_air_output.append(C4.air_output(control_input, air_demand, active_compressors)) #appends C4 air output reading to C4_air_output list
        C5_power_output.append(C5.power_output(control_input, air_demand, active_compressors)) #appends C5 power output reading to C5_power_output list
        C5_air_output.append(C5.air_output(control_input, air_demand, active_compressors)) #appends C5 air output reading to C5_air_output list
        C6_power_output.append(C6.power_output(control_input, air_demand, active_compressors)) #appends C6 power output reading to C6_power_output list
        C6_air_output.append(C6.air_output(control_input, air_demand, active_compressors)) #appends C6 air output reading to C6_air_output list
        control_input = False #control input for compressor.power_output and compressor.air_output
        
Control_Scheme_Results = openpyxl.load_workbook('cs3_op4_2019_results.xlsx') #opens results workbook in current directory
Compressors_Power_Output = Control_Scheme_Results['compressors_power_output'] #calls sheet 'compressors_power_output' in workbook
Compressors_Power_Output['D1'].value = 'C1_power_output' #names cell in workbook
for i in range(len(C1_power_output)):
    Compressors_Power_Output['D' + str(i + 2)].value = C1_power_output[i] #writes data in C1_power_output to sheet
Compressors_Power_Output['E1'].value = 'C2_power_output' #names cell in workbook
for i in range(len(C2_power_output)):
    Compressors_Power_Output['E' + str(i + 2)].value = C2_power_output[i] #writes data in C2_power_output to sheet
Compressors_Power_Output['F1'].value = 'C3_power_output' #names cell in workbook
for i in range(len(C3_power_output)):
    Compressors_Power_Output['F' + str(i + 2)].value = C3_power_output[i] #writes data in C3_power_output to sheet
Compressors_Power_Output['G1'].value = 'C4_power_output' #names cell in workbook
for i in range(len(C4_power_output)):
    Compressors_Power_Output['G' + str(i + 2)].value = C4_power_output[i] #writes data in C4_power_output to sheet
Compressors_Power_Output['H1'].value = 'C5_power_output' #names cell in workbook
for i in range(len(C5_power_output)):
    Compressors_Power_Output['H' + str(i + 2)].value = C5_power_output[i] #writes data in C5_power_output to sheet
Compressors_Power_Output['I1'].value = 'C6_power_output' #names cell in workbook
for i in range(len(C6_power_output)):
    Compressors_Power_Output['I' + str(i + 2)].value = C6_power_output[i] #writes data in C6_power_output to sheet

Compressors_Air_Output = Control_Scheme_Results['compressors_air_output'] #calls sheet 'compressors_air_output' in workbook
Compressors_Air_Output['D1'].value = 'C1_air_output' #names cell in workbook
for i in range(len(C1_air_output)):
    Compressors_Air_Output['D' + str(i + 2)].value = C1_air_output[i] #writes data in C1_air_output to sheet
Compressors_Air_Output['E1'].value = 'C2_air_output' #names cell in workbook
for i in range(len(C2_air_output)):
    Compressors_Air_Output['E' + str(i + 2)].value = C2_air_output[i] #writes data in C2_air_output to sheet
Compressors_Air_Output['F1'].value = 'C3_air_output' #names cell in workbook
for i in range(len(C3_air_output)):
    Compressors_Air_Output['F' + str(i + 2)].value = C3_air_output[i] #writes data in C3_air_output to sheet
Compressors_Air_Output['G1'].value = 'C4_air_output' #names cell in workbook
for i in range(len(C4_air_output)):
    4Compressors_Air_Output['G' + str(i + 2)].value = C4_air_output[i] #writes data in C4_air_output to sheet
Compressors_Air_Output['H1'].value = 'C5_air_output' #names cell in workbook
for i in range(len(C5_air_output)):
    Compressors_Air_Output['H' + str(i + 2)].value = C5_air_output[i] #writes data in C5_air_output to sheet
Compressors_Air_Output['I1'].value = 'C6_air_output' #names cell in workbook
for i in range(len(C6_air_output)):
    Compressors_Air_Output['I' + str(i + 2)].value = C6_air_output[i] #writes data in C6_air_output to sheet
   
Control_Scheme_Results.save('cs3_op4_2019_results.xlsx') #saves results workbook
Control_Scheme_Results.close() #closes results workbook
