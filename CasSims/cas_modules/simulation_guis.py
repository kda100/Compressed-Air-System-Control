# -*- coding: utf-8 -*-
"""
Created on Wed Sep 16 13:38:35 2020

@author: kda_1
"""

import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
import os
import openpyxl
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import threading

interface_title = "CasSims" #title of CasSims program

program_description = "CasSim can be used to perform simulations on your compressed air system.\n\nIt gives you detailed performance metrics on the behaviours of your compressors."

class MainInterface():
    '''
    Creates main interface used to introduce and explains the compressed air system simulation program.
    '''
    def __init__(self, master, main_function):
        self._master = master #parent window
        self._master.title(interface_title)
        self._master.geometry("600x400")
        self._program_title = tk.Label(self._master, 
                                      text=f"Welcome to the {interface_title} Program", 
                                      font=("Arial", 20), 
                                      fg="blue4"
                                      ) #label displaying title of program
        self._program_title.grid(row=0, column=0, pady=10, padx=10, columnspan=3)
        
        self._program_description = tk.Label(self._master, 
                                            text=program_description, 
                                            font=("Time New Roman", 10), 
                                            fg="black"
                                            ) #label with a discription of program
        self._program_description.grid(row=1, column=0, pady=10, padx=10, columnspan=3)
        
        self._num_simulations_spinbox_label = tk.Label(self._master, 
                                                      text="Number of simulations to be performed: ",
                                                      font=("Arial", 12), 
                                                      fg="black", 
                                                      justify='left'
                                                      ) #spinbox title label
        self._num_simulations_spinbox_label.grid(row=2, column=0, pady=10, padx=10)
        
        self._num_simulations_spinbox = tk.Spinbox(self._master, 
                                                   values=["1"], 
                                                   font=("Arial", 12)
                                                   ) #spinbox widget to store number of simulations
        self._num_simulations_spinbox.grid(row=2, column=1, pady=10, padx=10)
        
        self._start_button = tk.Button(self._master, text="Start", #button to start getting details for simulations
                                      command=self.start_get_simulation_info, 
                                      height=2, width=8)
        self._start_button.grid(row=2, column=2, pady=10, padx=10)
        
        self._main_function = main_function #main function to be done at the end
        self._master.protocol("WM_DELETE_WINDOW", self.on_exit) #changes "X" button functionally on interface
        
        for row in range(3): 
            self._master.grid_rowconfigure(row, weight=1) #configures rows in master
        for col in range(3):
            self._master.grid_columnconfigure(col, weight=1) #configures columns in master
    
    @property
    def master(self):
        return self._master
    
    @property
    def start_button(self):
        return self._start_button
    
    def on_exit(self):
        '''
        Creates a popup message if "x" is clicked on MainInterface to exit program.
        '''
        if messagebox.askyesno("Exit", 
                               "Are you sure you want to quit the program?", 
                               parent=self._master):
            self._master.destroy()
    
    
    def start_get_simulation_info(self):
        '''
        Starts next stage of program by asking user for the details on the compressors and the control sequence
        of the simulation they would like to perform.
        '''
        self.num_of_simulations = int(self._num_simulations_spinbox.get()) #gets input from the spinbox
        self._start_button['state'] = tk.DISABLED #prevents user from starting program twice
        self._main_function() #call get simulation details function
    
    
        
class SimulationUserInputsInterface():
    '''
    Creates an interface that prompts user to input the details of the control 
    '''
    def __init__ (self, master, num_of_simulations, main, main_function):
        
        
        self._master = master
        self._master.title(interface_title) #changes properties of interface
        self._master.geometry("800x600")
        
        self._num_of_compressors_frame = tk.Frame(self._master) #frame in interface to get number of compressors
        self._num_of_compressors_frame.pack()
        
        self._compressor_inputs_title_frame = tk.Frame(self._master) #frame in interface to store title compressor inputs
        self._compressor_inputs_title_frame.pack()
        
        self._compressor_inputs_frame = tk.Frame(self._master) #frames in interface to store entry widgets for compressor inputs
        self._compressor_inputs_frame.pack()
        
        self._num_compressors_spinbox_label = tk.Label(self._num_of_compressors_frame, 
                                                       text="Number of compressors in simulation: ", 
                                                       font=("Arial", 12), 
                                                       fg="black", 
                                                       justify='left' 
                                                       ) #label title for number of compressors spinbox
        self._num_compressors_spinbox_label.grid(row=0, column=0, pady=10, padx=10, columnspan=2, sticky=tk.N)
        
        self._num_compressors_spinbox = tk.Spinbox(self._num_of_compressors_frame, 
                                                   from_=1, to=5, font=("Arial", 12), 
                                                   width=30
                                                   ) #spinbox for user input of number of compressors.
        self._num_compressors_spinbox.grid(row=0, column=2, pady=10, padx=10, sticky=tk.N)
        
        self._get_compressor_details_button = tk.Button(self._num_of_compressors_frame, text="Next", 
                                                          command=self.get_compressors, height=1, width=8
                                                          )# button to get entries of max power, max air flow and operation mode of compressors
        self._get_compressor_details_button.grid(row=0, column=3, pady=10, padx=10, sticky=tk.N)
        
        self._num_of_compressors_frame.grid_rowconfigure(0, weight=1)
        for col in range(3):
            self._num_of_compressors_frame.grid_columnconfigure(col, weight=1) #configures grids in self._num_of_compressors_frame
            
        self._master.protocol("WM_DELETE_WINDOW", self.on_exit) #changes "X" button functionally on interface
        self._main=main
        
        self._main_function = main_function #main function to be done at the end #function once all simulation details have been inputted
        
        self._trim_topLevel = None #assigns toplevels for different different user inputs
        self._control_sequence_type_topLevel = None
        self._setpoints_num_topLevel = None
        self._assign_setpoints_topLevel = None
        self._upload_sheet_topLevel = None
        
        menu = tk.Menu(self._master) #add menu to master/toplevel
        self._master.config(menu=menu)
        file_menu=tk.Menu(menu, tearoff=False)
        menu.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Restart", command=self.restart_program) #adds restart program command/button to menu >> Restart
        file_menu.add_command(label="Exit", command=self.exit_program) #add exit program command/button to menu >> File
        self._master.lift(main.master) #lifts interfaces above main interface
        
    def restart_program(self):
        '''
        Creates popup message to close SimulationUserInputsInterface and enabled start button in MainInterface for user
        to restart simulation inputs.
        '''
        if messagebox.askyesno("Exit", "Are you sure you want to restart the simulation?", parent=self._master):
            self._master.destroy()
            self._main.start_button['state'] = tk.NORMAL #enables start button on main
            if self._trim_topLevel != None: #checks if other toplevels exists and then closes them
                self._trim_topLevel.destroy()
            if self._control_sequence_type_topLevel != None:
                self._control_sequence_type_topLevel.destroy()
            if self._setpoints_num_topLevel != None:
                self._setpoints_num_topLevel.destroy()
            if self._assign_setpoints_topLevel != None:
                self._assign_setpoints_topLevel.destroy()
            if self._upload_sheet_topLevel != None:
                self._upload_sheet_topLevel.destroy()
    
    def exit_program(self):
        '''
        Creates a popup message to close entire program.
        '''
        if messagebox.askyesno("Exit", "Are you sure you want to quit the program?", parent = self._main.master):
            self._main.master.destroy() #closes 
            
    def on_exit(self):
        '''
        Creates a popup message if "x" is clicked on master.
        '''
        if messagebox.askyesno("Exit", "Are you sure you want to quit the simulation?", parent=self._master):
            self._master.destroy()
            self._main.start_button['state'] = tk.NORMAL
            if self._trim_topLevel != None: #checks if other toplevels exists and then closes them
                self._trim_topLevel.destroy()
            if self._control_sequence_type_topLevel != None:
                self._control_sequence_type_topLevel.destroy()
            if self._setpoints_num_topLevel != None:
                self._setpoints_num_topLevel.destroy()
            if self._assign_setpoints_topLevel != None:
                self._assign_setpoints_topLevel.destroy()
            if self._upload_sheet_topLevel != None:
                self._upload_sheet_topLevel.destroy()    
    
    def on_exit_top(self, topLevel, button):
            '''
            Enables button if 'x' button is clicked on topLevel interface.
            '''
            topLevel.destroy()
            button['state'] = tk.NORMAL
    
    def get_compressors(self):
        
        '''
        Creates entry widgets and labels for user to input maximum power, maximum air flow and operation modes of compressors
        '''
        
        self.num_of_compressors = int(self._num_compressors_spinbox.get()) #gets num of compressors from spinbox
        self._get_compressor_details_button["state"] = "disabled" #prevents user from getting compressors details again
        
        compressor_details_title = tk.Label(self._compressor_inputs_title_frame, font=("Arial", 16), 
                                            text = "Choose the maximum power(s), maximum air flow(s) and operation mode(s)\nfor the compressor(s) in the simulation")
        compressor_details_title.grid(row=0, column=0, sticky=tk.NSEW, pady=10, padx=10)
        
        self._compressor_inputs_title_frame.grid_columnconfigure(0, weight=1) #configures grid in frame
        self._compressor_inputs_title_frame.grid_rowconfigure(0, weight=1)
        
        compressors_label = tk.Label(self._compressor_inputs_frame, text="Compressors", font=("Helvetica", 10))
        compressors_label.grid(row=0, column=0, pady=10, padx=10, sticky=tk.N) #creates column title for compressors
        
        maximum_power_label = tk.Label(self._compressor_inputs_frame, text="Maximum Power (kW)", font=("Helvetica", 10))
        maximum_power_label.grid(row=0, column=1, pady=10, padx=10, sticky=tk.N) #creates column titles for max power of compressors
        
        maximum_air_flow_label = tk.Label(self._compressor_inputs_frame, text="Maximum Air Flow (m3/hr)", font=("Helvetica", 10))
        maximum_air_flow_label.grid(row=0, column=2, pady=10, padx=10, sticky=tk.N) #creates column titles for max air flow of compressors
        
        operation_mode_label = tk.Label(self._compressor_inputs_frame, text="Operation Mode", font=("Helvetica", 10))
        operation_mode_label.grid(row=0, column=3, pady=10, padx=10,sticky=tk.N) #creates column titles for operation modes of compressors
        
        for row in range(1): #configures rows and columns in _compressor_inputs_title_frame
            self._compressor_inputs_title_frame.grid_rowconfigure(row, weight=1)
        for column in range(4):
            self._compressor_inputs_title_frame.grid_columnconfigure(column, weight=1)
        
        #creates dictionaries for maximum_power, maximum_air_flow and operation_mode entries
        self.maximum_power_entries = {}
        self.maximum_air_flow_entries = {}
        self.operation_mode_entries = {}
        for i in range(self.num_of_compressors):  #for each compressor stores an empty list in dictionary for entry and value of entry
            self.maximum_power_entries[f'C{i+1}'] = []
            self.maximum_air_flow_entries[f'C{i+1}'] = []
            self.operation_mode_entries[f'C{i+1}'] = []
            
        for i in range(self.num_of_compressors):
            self._compressor_inputs_frame.grid_rowconfigure(i+1, weight=1) #configures row for each row of compressor inputs
            for j in range(4):
                if j == 0: #for first column in _compressor_inputs_title_frame
                    compressor_label = tk.Label(self._compressor_inputs_frame, text=f"C{i+1}") #creates label for each compressor 
                    compressor_label.grid(row=i+1, column=j, pady=5, padx=5, sticky=tk.N)
                    
                elif j == 1: #puts entry widget for max power in second column and store entry widget in maximum_power_entries
                    maximum_power_entry = tk.Entry(self._compressor_inputs_frame)
                    maximum_power_entry.grid(row=i+1, column=j, sticky=tk.N, pady=5, padx=10)
                    self.maximum_power_entries[f'C{i+1}'].append(maximum_power_entry)
                    
                elif j == 2: #puts entry widget for max air flow in third column and store entry widget in maximum_air_flow_entry
                    maximum_air_flow_entry = tk.Entry(self._compressor_inputs_frame)
                    maximum_air_flow_entry.grid(row=i+1, column=j, sticky=tk.N, pady=5, padx=10)
                    self.maximum_air_flow_entries[f'C{i+1}'].append(maximum_air_flow_entry)
                    
                elif j == 3: #puts OptionMenu for operation mode in fourth column and stores option in operation_mode_entries
                    COMPRESSOR_TYPES = ["OnOff", "LoadUnload", "InletModulation", "VariableSpeed"]
                    compressor_type = tk.StringVar()
                    compressor_type.set(COMPRESSOR_TYPES[0]) #sets initial value of option menu
                    compressor_type_entry = tk.OptionMenu(self._compressor_inputs_frame, compressor_type, *COMPRESSOR_TYPES)
                    compressor_type_entry.grid(row=i+1, column=j, sticky=tk.N, pady=5, padx=10)
                    self.operation_mode_entries[f'C{i+1}'].append(compressor_type)

        self.check_compressor_entries_button = tk.Button(self._compressor_inputs_frame,  #creates button to ask user for trim compressor
                                                    text="Next", command=self.check_compressor_entries, 
                                                    height=1, width=8)
        self.check_compressor_entries_button.grid(row=self.num_of_compressors+1, column=3, padx=10, pady=10, sticky=tk.N)
        self.check_compressor_entries_button.grid_rowconfigure(self.num_of_compressors+1, weight=1)
    
    
    def check_compressor_entries(self):
        '''
        Checks if maximum power and maximum air flow entries of compressors are valid then saves them
        '''
        try:
            for i in range(self.num_of_compressors): #deletes any existing results of entries from max power, max air flow and operation mode dicts
                if len(self.maximum_power_entries[f'C{i+1}'])>1:
                    del self.maximum_power_entries[f'C{i+1}'][-1]
                elif len(self.maximum_air_flow_entries[f'C{i+1}'])>1:
                    del self.maximum_air_flow_entries[f'C{i+1}'][-1]
                elif len(self.operation_mode_entries[f'C{i+1}'])>1:
                    del self.operation_mode_entries[f'C{i+1}'][-1]
            
            for i in range(self.num_of_compressors): #stores entries of compressor details into dictionary
                self.maximum_power_entries[f'C{i+1}'].append(float(self.maximum_power_entries[f'C{i+1}'][0].get())) #ensures entry is a float
                self.maximum_air_flow_entries[f'C{i+1}'].append(float(self.maximum_air_flow_entries[f'C{i+1}'][0].get())) #ensures entry is a float
                self.operation_mode_entries[f'C{i+1}'].append(self.operation_mode_entries[f'C{i+1}'][0].get())
            
            total_compressor_max_air_flow = 0
            for i in range(self.num_of_compressors): #calculates the compressors and max air flow 
                total_compressor_max_air_flow += self.maximum_air_flow_entries[f'C{i+1}'][-1]
            if total_compressor_max_air_flow> 10000:
                raise MemoryError #checks if maximum air flow of compressors is too great
        except ValueError: #if entry is not a number
            messagebox.showwarning("Input Error", "Maximum Power and Maximum Air Flow Input(s)\nmust be a number.", parent=self._master)
        except MemoryError:
            messagebox.showwarning("Input Error", "Total Compressor Maximum Air Flow is too great!", parent=self._master)
        else: #disables entry widgets if entries are okay
            for i in range(self.num_of_compressors):
                self.maximum_power_entries[f'C{i+1}'][0].config(state=tk.DISABLED)
                self.maximum_air_flow_entries[f'C{i+1}'][0].config(state=tk.DISABLED)
            
            self.get_trim_compressor() #asks user for trim compressor
    
    
    def get_trim_compressor(self):
        '''
        Assigns trim compressor for user if only one compressor in simulation or creates an
        interface to ask user for trim compressor
        '''
        
        self.check_compressor_entries_button['state']=tk.DISABLED #prevents user from using button
                
        if self.num_of_compressors==1: #if only one compressor chosen trim is set as first compressor
            self.trim = 'C1'
            one_trim_label = tk.Label(self._compressor_inputs_frame, 
                                      text="Only one compressor chosen so trim is 'C1'", 
                                      font=("Helvetica", 16), fg="red") #label is shown on interface that one compressor was chosen 
            one_trim_label.grid(row=self.num_of_compressors+1, column=0, columnspan=3, padx=20, sticky=tk.N)
            self.get_control_scheme_type() #then user is prompted to pick control scheme type
            
        else: #user chooses more than one compressor.
            self._trim_topLevel = tk.Toplevel() #creates a interface to ask user for trim compressor
            self._trim_topLevel.title(interface_title)
            self._trim_topLevel.geometry("600x75")
            self._trim_topLevel.protocol("WM_DELETE_WINDOW", lambda:self.on_exit_top(self._trim_topLevel, 
                                                                                     self.check_compressor_entries_button)) #changes "X" button functionally on interface
                
    
            ask_trim_label = tk.Label(self._trim_topLevel, 
                                  text="Choose the trim compressor for the simulation: ") #creates label to ask user for trim compresor
            ask_trim_label.grid(row=0, column=0, pady=10, padx=10, sticky=tk.N)
            
            COMPRESSORS = [f'C{x+1}' for x in range(self.num_of_compressors)] #creates list for trim compressor options
            trim_compressor = tk.StringVar() #variable to store user compressor input
            trim_compressor.set(COMPRESSORS[0])
            ask_trim_entry = tk.OptionMenu(self._trim_topLevel, trim_compressor, *COMPRESSORS) #creates option menu to ask user for trim compressor
            ask_trim_entry.config(width=8)
            ask_trim_entry.grid(row=0, column=1, padx=10, pady=10, sticky=tk.N)
            
            
            def save_trim_compressor():
                '''
                Saves trim compressor chosen by user then prompts users to pick the control scheme type
                '''
                
                self.trim = trim_compressor.get() #gets user input of trim compressor
                self._trim_topLevel.destroy()  #closes interface
                self.get_control_scheme_type() #prompts user to input control scheme type.
            
            
            store_trim_compressor_button = tk.Button(self._trim_topLevel, text="Next", 
                                                       command=save_trim_compressor, 
                                                       height=1, width=6) #creates button to save user input of trim compressor and prompt user to pick control scheme type.
            store_trim_compressor_button.grid(row=0, column=2, padx=10, pady=10, sticky=tk.N)
            
            self._trim_topLevel.grid_rowconfigure(0, weight=1) #configs grids on _trim_topLevel interface.
            for col in range(3):
                self._trim_topLevel.grid_columnconfigure(col, weight=1)
            
            self._trim_topLevel.lift(self._master)
        
    def get_control_scheme_type(self):
        '''
        Prompts user to choose the control scheme type to be performed on their simulation
        '''
        self._control_sequence_type_topLevel = tk.Toplevel() #creates interface 
        self._control_sequence_type_topLevel.title(interface_title)
        self._control_sequence_type_topLevel.geometry("600x75")
        self._control_sequence_type_topLevel.protocol("WM_DELETE_WINDOW", lambda:self.on_exit_top(self._control_sequence_type_topLevel,
                                                                                                self.check_compressor_entries_button)) #changes "X" button functionally on interface
        
        
        control_type_label = tk.Label(self._control_sequence_type_topLevel, 
                                      text="Choose the control scheme type for the simulation: ") #labels to instruct user to choose control type
        control_type_label.grid(row=0, column=0, padx=20, pady=20)
        
        CONTROL_SEQUENCE_TYPES = ["Optimisation", "Defined"] #list of control types available
        control_sequence_type = tk.StringVar() #variable to store control_scheme_type
        control_sequence_type.set(CONTROL_SEQUENCE_TYPES[0])
        control_sequence_type_entry = tk.OptionMenu(self._control_sequence_type_topLevel, 
                                           control_sequence_type, 
                                           *CONTROL_SEQUENCE_TYPES) #creates option menu for user to choose control scheme type
        control_sequence_type_entry.grid(row=0, column=1, padx=20, pady=20, sticky=tk.N)
        
        
        def save_control_scheme_type():
            '''
            Saves control scheme type and prompts user to either upload an excel workbook or choose setpoints
            '''
            CONTROL_SEQUENCE_TYPES2 = {"Optimisation":"opt", "Defined":"def"}
            self.control_sequence_type = CONTROL_SEQUENCE_TYPES2[control_sequence_type.get()] #gets user input of trim compressor
            self._control_sequence_type_topLevel.destroy()  #closes interface
            
            if self.control_sequence_type == "opt":
                self.create_choose_workbook_button()
            if self.control_sequence_type == "def":
                self.get_setpoints()
        
        get_control_type_button = tk.Button(self._control_sequence_type_topLevel, text="Next", 
                                            command=save_control_scheme_type, 
                                            height=1, width=6) #creates button for user to save control scheme type
        get_control_type_button.grid(row=0, column=2, padx=20, pady=20, sticky=tk.N)
        
        self._control_sequence_type_topLevel.grid_rowconfigure(0, weight=1) #configures grids in interface
        for col in range(3):
            self._control_sequence_type_topLevel.grid_columnconfigure(col, weight=1)
        self._control_sequence_type_topLevel.lift(self._master)
        
        
    def get_setpoints(self):
        '''
        Prompts user to choose setpoints or automatically assigns the first compressor's maximum power as the setpoint
        '''
        
        if self.num_of_compressors==1: #only one compressor in simulation
            one_setpoint_label = tk.Label(self._compressor_inputs_frame, 
                                          text="Only one compressor chosen so only one setpoint", 
                                          font=("Helvetica", 16), fg="red") #tells user that one compressor is chosen so only one setpoint
            one_setpoint_label.grid(row=self.num_of_compressors+2, column=0, columnspan=3, padx=20, sticky=tk.W)
            self.create_choose_workbook_button()
            self.num_of_setpoints = 1
            
        elif self.num_of_compressors>1: #more than one compressor in simulation
            self._setpoints_button = tk.Button(self._compressor_inputs_frame, 
                                                     text="Choose Setpoints", 
                                                     command=self.get_setpoint_num) #creates button to get the number of setpoints on master
            self._setpoints_button.grid(row=self.num_of_compressors+2, column=3, padx=10, pady=10)
            self._compressor_inputs_frame.grid_rowconfigure(self.num_of_compressors+2, weight=1) #configures the row

        self._master.lift(self._main.master) #lifts master above main interface


    def get_setpoint_num(self):
        '''
        Prompts user to input the number of setpoints for the control scheme in their simulation
        '''
        
        self._setpoints_button['state'] = tk.DISABLED #disabled button so user cannot click twice
        self._setpoints_num_topLevel = tk.Toplevel() #creates interface for user to choose num of setpoints
        self._setpoints_num_topLevel.title(interface_title) 
        self._setpoints_num_topLevel.geometry("600x75") 
        self._setpoints_num_topLevel.protocol("WM_DELETE_WINDOW", lambda:self.on_exit_top(self._setpoints_num_topLevel,
                                                                                          self._setpoints_button)) #changes "X" button functionally on interface
            
        num_of_setpoints_spinbox_label = tk.Label(self._setpoints_num_topLevel, 
                                      text="Choose Number of Setpoints: ", 
                                      font=("Arial", 12), fg="black", 
                                      justify='left') #label title for spinbox for user to choose number of setpoints
        num_of_setpoints_spinbox_label.grid(row=0, column=0, pady=10, padx=10, sticky=tk.N)
        
        num_of_setpoints_spinbox = tk.Spinbox(self._setpoints_num_topLevel, 
                                                   from_=2, to=12, font=("Arial", 12)) #spinbox for user to choose number of setpoints
        num_of_setpoints_spinbox.grid(row=0, column=1, pady=10, padx=10, sticky=tk.N)
        
        self._setpoints_num_topLevel.grid_rowconfigure(0, weight=1) #configures grids in _setpoints_num_topLevel
        for col in range(3):
            self._setpoints_num_topLevel.grid_columnconfigure(col, weight=1)
        self._setpoints_num_topLevel.lift(self._master)
        
        
        def save_num_of_setpoints():
            '''
            Saves number of setpoints then prompts user to assign compressors to setpoints
            '''
            self.num_of_setpoints = int(num_of_setpoints_spinbox.get())
            self._setpoints_num_topLevel.destroy()
            self.assign_setpoints()
            
            
        assign_setpoints_button = tk.Button(self._setpoints_num_topLevel, text="Next", 
                                            command=save_num_of_setpoints, 
                                            height=1, width=6)
        assign_setpoints_button.grid(row=0, column=3, pady=10, padx=10, sticky=tk.N) #creates button to save number of setpoints
        
    def assign_setpoints(self):
        '''
        Creates interface for user to assign compressors to setpoints
        '''
        self._assign_setpoints_topLevel = tk.Toplevel() #creates interface to assign compressors to setpoints
        self._assign_setpoints_topLevel.title("Compressed Air System Simulation(s)")
        self._assign_setpoints_topLevel.geometry("400x500")
        self._assign_setpoints_topLevel.protocol("WM_DELETE_WINDOW", lambda:self.on_exit_top(self._assign_setpoints_topLevel,
                                                                                             self._setpoints_button)) #changes "X" button functionally on interface
        
        assign_setpoints_title_frame = tk.Frame(self._assign_setpoints_topLevel) #creates frame for assign_setpoints_title
        assign_setpoints_title_frame.pack()
        assign_setpoints_title = tk.Label(assign_setpoints_title_frame, 
                                          text="Choose the what compressor(s) maximum capacities\nare contributing to each setpoint:", 
                                          font=("Arial", 12), fg='blue4') #creates title for choosing setpoints interface
        assign_setpoints_title.grid(row=0, column=0, columnspan=2, pady=10, padx=10, sticky=tk.N)
        
        assign_setpoints_title_frame.grid_rowconfigure(0, weight = 1)
        for i in range(2): #configures grids in assign_setpoints_title_frame
            assign_setpoints_title_frame.grid_columnconfigure(i+1, weight = 1)
        
        self.setpoint_entries = {} #creates a dictionary to store setpoint entry widgets and the value of the entries
        for i in range(self.num_of_setpoints):
            self.setpoint_entries[f's{i+1}'] = [] #stores a list for each setpoint
        
        setpoint_entries_frame = tk.Frame(self._assign_setpoints_topLevel) #creates a frame for setpoint entries
        setpoint_entries_frame.pack()
        
        for i in range(self.num_of_setpoints):
            for j in range(2):
                if j == 0: #creates a label for each setpoint entry widget
                    setpoint_label = tk.Label(setpoint_entries_frame, text=f"Setpoint {i+1}")
                    setpoint_label.grid(row=i+1, column=j, pady=5, padx=10, sticky=tk.N)
                if j == 1: #creates a entry widget for each setpoint
                    setpoint_entry = tk.Entry(setpoint_entries_frame)
                    setpoint_entry.grid(row=i+1, column=j, pady=5, padx=10, sticky=tk.N)
                    self.setpoint_entries[f's{i+1}'].append(setpoint_entry) #adds entry widget to dictionary of setpoint for each setpoint
                    
        check_setpoint_entries_button = tk.Button(setpoint_entries_frame, 
                                                  text="Next", 
                                                  command=self.check_setpoint_entries)
        check_setpoint_entries_button.grid(row=self.num_of_setpoints+1, column=1, padx=10, pady=10, sticky=tk.N)
        
        self._assign_setpoints_topLevel.lift(self._master)#lifts _assign_setpoints_topLevel interface
        
        for i in range(self.num_of_setpoints+1): #configures the rows in setpoint_entries_frame
            setpoint_entries_frame.grid_rowconfigure(i, weight = 1)
        for i in range(2): #configures columns in setpoint_entries_frame
            setpoint_entries_frame.grid_columnconfigure(i+1, weight = 1)
    
    def check_setpoint_entries(self):
        '''
        Checks if the entries of the setpoints are valid
        '''
        for i in range(self.num_of_setpoints):
                if len(self.setpoint_entries[f's{i+1}'])>1:
                    del self.setpoint_entries[f's{i+1}'][-1] #deletes any setpoint entries that exist
        try:
            for i in range(self.num_of_setpoints):
                setpoint_compressors = [x.strip() for x in self.setpoint_entries[f's{i+1}'][0].get().upper().split(',')] #checks if compressor objects are separated by commas
                for compressor in setpoint_compressors:
                    if compressor not in [f'C{x+1}' for x in range(self.num_of_compressors)]: #checks if compressors choosen are actual compressors in the simulation
                        raise NameError
        except:
            messagebox.showwarning("Input Error", f'Compressor(s) chosen must be at least one of the compressors C1 - C{self.num_of_compressors}, multiple compressors should be written with a "," between them\n', 
                                   parent=self._assign_setpoints_topLevel) #creates an error message
        else:
             for i in range(self.num_of_setpoints):
                 self.setpoint_entries[f's{i+1}'].append([x.strip() for x in self.setpoint_entries[f's{i+1}'][0].get().upper().split(',')]) #creates a list for compressors in entries for each setpoint
                 
             self._assign_setpoints_topLevel.destroy() #destroys interface
             self.create_choose_workbook_button()
       
        
    def create_choose_workbook_button(self):
        '''
        Creates button for user to choose workbook with compressed air demand data for simulation
        '''
        choose_workbook_label = tk.Label(self._compressor_inputs_frame, 
                                         text="Choose Excel Workbook with\nAir Demand Data: ") #creates a label for choosing excel workbook button
        choose_workbook_label.grid(row=self.num_of_compressors+3, column=3, pady=(10,0), padx=10)
        
        self.choose_workbook_button = tk.Button(self._compressor_inputs_frame, text="Upload File", 
                                                command=self.get_workbook, height=1, width=10) #creates button to open a filedialog for user to choose workbook
        self.choose_workbook_button.grid(row=self.num_of_compressors+4, column=3, padx=10)
        
    
    def get_workbook(self):
        '''
        Prompts user to choose location of excel workbook
        '''
        self.choose_workbook_button['state'] = tk.DISABLED #prevent user from choosing workbook twice
        excel_filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select Excel Workbook", 
                                                         filetypes=[("Excel files", "*.xlsx")], parent=self._master)
        
        t = threading.Thread(target=lambda:self.upload_workbook(excel_filename)) #creates a thread to upload the workbook
        t.start()
    
    
    def upload_workbook(self, excel_filename):
        '''
        Uploads the workbook chosen by the user.
        '''
        try:
            self.workbook = openpyxl.load_workbook(f"{excel_filename}") #uploads excel file
        except PermissionError: #if excel file is still open
            messagebox.showerror("Workbook Error", "Ensure Workbook is closed", parent=self._master)
            self.choose_workbook_button['state'] = tk.NORMAL #allows user to choose workbook again
        except:
            self.choose_workbook_button['state'] = tk.NORMAL #allows user to choose workbook again
        else:
            self.choose_worksheet()
            
    
    def choose_worksheet(self):
        '''
        Creates interface for user to choose excel worksheet with air demand data.
        '''
        self._upload_sheet_topLevel = tk.Toplevel() #creates interface 
        self._upload_sheet_topLevel.title("Compressed Air System Simulation(s)")
        self._upload_sheet_topLevel.geometry("400x400")
        self._upload_sheet_topLevel.protocol("WM_DELETE_WINDOW", lambda:self.on_exit_top(self._upload_sheet_topLevel,
                                                                                             self.choose_workbook_button)) #changes "X" button functionally on interface
            
        _upload_sheet_topLevel_title = tk.Label(self._upload_sheet_topLevel,
                                                text="Choose Worksheet with Air Demand Data:", 
                                                font=("Arial", 12), fg='blue4') #creates title for interface
        _upload_sheet_topLevel_title.pack(anchor="w", pady=20, padx=20)
        
        sheet = tk.StringVar() #creates a variable for excel sheet
        sheet.set(self.workbook.sheetnames[0]) #sets option to be first sheet in list
        for sheet_name in self.workbook.sheetnames:
            sheet_button = tk.Radiobutton(self._upload_sheet_topLevel, text=sheet_name, variable=sheet, value=sheet_name)
            sheet_button.pack(anchor="w", padx=10, pady=5) #creates radio button for each sheet
        
        self.sheet = sheet #saves sheet
        
        get_workbook_button = tk.Button(self._upload_sheet_topLevel, text="Upload Sheet",
                                        command=self.create_check_worksheet_thread) #creates button to check if data in worksheet is valid
        get_workbook_button.pack(anchor="w", pady=10, padx=10)
        self._upload_sheet_topLevel.lift(self._master) #lifts_upload_sheet_topLevel above master
            
    def create_check_worksheet_thread(self):
        '''
        Creates thread to check worksheet
        '''
        check_worksheet_thread = threading.Thread(target=self.check_worksheet)
        check_worksheet_thread.start()
        
    def check_worksheet(self):
        '''
        Checks if air demand data in worksheet is valid
        '''
        self.air_demand_data_sheet = self.workbook[self.sheet.get()] #svaes sheetname
        try:
            for row in self.air_demand_data_sheet.iter_rows(min_row = 1, max_row = 1, min_col = 1, max_col = self.air_demand_data_sheet.max_column): #iterates over first row of air_volume_stacked, excel sheet object
                col = 1
                for cell in row: #iterates over each cell in row
                    col = col + 1
                    if cell.value == 'CA_READINGS': #checks for cell in first row that contains 'CA_READINGS'
                        self.air_demand_data_column = cell.coordinate[0] #stores the coordinate of cell with 'CA_READINGS'
                    elif col == self.air_demand_data_sheet.max_column:
                        raise NameError
            
            max_air_demand = 0
            for cell in self.air_demand_data_sheet[self.air_demand_data_column]: #checks the maximum air_demand in sheet
                if type(cell.value) == int or type(cell.value) == float:
                    if cell.value > max_air_demand:
                        max_air_demand=cell.value 
                        
            compressors_maximum_air_flow = 0 #calculates the maximum air flow of compressors choosen
            for i in range(self.num_of_compressors):
                compressors_maximum_air_flow = compressors_maximum_air_flow + self.maximum_air_flow_entries[f'C{i+1}'][-1] 
                
            if max_air_demand>compressors_maximum_air_flow: #compressors incapable of suppling air demand in simulation
                raise ValueError
                
        except NameError:
            messagebox.showerror("Invalid Sheet", "Ensure 'CA_reading' is in first row of column\nwith the compressed air volume data", 
                                 parent = self._upload_sheet_topLevel) #message box for 'CA_READINGS' not being seen
        except ValueError:
            messagebox.showerror("Insufficient Compressor Capacity", "Compressors in simulation have insufficient capacity (m3/hr)\nto perform simulation, please restart program", 
                                 parent = self._upload_sheet_topLevel) #message box for compressors having insuficient capacity
        else:
            self._upload_sheet_topLevel.destroy() #close _upload_sheet_topLevel interface
            self.create_perform_simulation_button() #creates new button on master
            
            
    def create_perform_simulation_button(self):
        '''
        Creates perform simulation button on master.
        '''
        
        perform_sim_button1 = tk.Button(self._compressor_inputs_frame, text="Perform Simulation", command=self.perform_sim)
        perform_sim_button1.grid(row=self.num_of_compressors+5, column=3, pady=20) #creates button to perform simulation
        
        self._compressor_inputs_frame.grid_rowconfigure(self.num_of_compressors+5, weight=1) #configures row for new button
        
        self._master.lift(self._main.master) #lifts master above main interface
            
    def perform_sim(self):
        self._master.destroy() #destroys master
        self._main_function() #runs function
             
class NameResultsInterface():
    def __init__(self, master, main, main_function):
        self._master = master #configures interface
        self._master.title(interface_title)
        self._master.geometry("600x75")
        
        self._get_results_name_label = tk.Label(master, text="Name the results: ") #title for entry widget to get results name
        self._get_results_name_label.grid(row=0, column=0, pady=10, padx=10, sticky=tk.N)
        
        self._get_results_name_entry =  tk.Entry(master) #entry widget to for user to input results name
        self._get_results_name_entry.grid(row=0, column=1, pady=10, padx=10, sticky=tk.N)
        
        self._get_results_name_button = tk.Button(master, text="Generate Results", command=self.check_results_name) #button for user to save results name       
        self._get_results_name_button.grid(row=0, column=2, pady=10, padx=10, sticky=tk.N)
        
        self._master.protocol("WM_DELETE_WINDOW", self.on_exit) #changes "X" button functionally on interface 
        self._main=main
        self._main_function=main_function #function once name has been inputted
        
        self._master.grid_rowconfigure(0, weight=1) #configures grid in interface
        for col in range(3):
            self._master.grid_columnconfigure(col, weight=1)
            
        menu = tk.Menu(self._master) #add menu to master/toplevel
        self._master.config(menu=menu)
        file_menu=tk.Menu(menu, tearoff=False)
        menu.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Restart", command=self.restart_program) #adds restart program command/button to menu >> Restart
        file_menu.add_command(label="Exit", command=self.exit_program) #add exit program command/button to menu >> File
        self._master.lift(self._main.master) #lifts interfaces above main interface
        
        
    def restart_program(self):
        '''
        Creates popup message to close SimulationUserInputsInterface and enabled start button in MainInterface for user
        to restart simulation inputs.
        '''
        if messagebox.askyesno("Exit", "Are you sure you want to restart the simulation?", 
                               parent = self._master):
            self._master.destroy() #destroys master
            self._main.start_button['state'] = tk.NORMAL
    
    
    def exit_program(self):
        '''
        Creates a popup message to close entire program.
        '''
        if messagebox.askyesno("Exit", "Are you sure you want to quit the program?", 
                               parent = self._main.master):
            self._main.master.destroy()
    
    
    def on_exit(self):
        '''
        Creates a popup message if "x" is clicked on master.
        '''
        if messagebox.askyesno("Exit", "Are you sure you want to quit the simulation?", 
                               parent = self._master):
            self._master.destroy()
            self._main.start_button['state'] = tk.NORMAL
    
    def check_results_name(self):
        '''
        Checks results name entry is valid
        '''
        try:
            self.results_name = self._get_results_name_entry.get()
            if len(self.results_name) == 0 or self.results_name.isspace() == True: #checks if no entry and or only whitespace
                raise NameError
        except:
            messagebox.showerror("Invalid Input", "Please Enter a Name.", parent=self._master)
        else:
            self.save_results_name()
    
    def save_results_name(self):
        '''
        Saves name from results_name_entry
        '''
        self.results_name = self.results_name.strip() #saves name
        self._master.destroy() #destroys master
        self._main_function() #displays results
            
class ShowResultsProgram():
    def __init__(self, master, fig, main):
        self._master = master #configures interface
        self._master.title(interface_title)
        
        self._fig = fig #fig containing results
        self._master.protocol("WM_DELETE_WINDOW", self.on_exit)
        self._main=main
        
        canvas = FigureCanvasTkAgg(self._fig, self._master) #creates canvas for results
        toolbar = NavigationToolbar2Tk(canvas, self._master)
        canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
        toolbar.update()
        
        menu = tk.Menu(master) #add menu to master/toplevel
        self._master.config(menu=menu)
        file_menu=tk.Menu(menu, tearoff=False)
        menu.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Restart", command=self.restart_program) #adds restart program command/button to menu >> Restart
        file_menu.add_command(label="Exit", command=self.exit_program) #add exit program command/button to menu >> File
        self._master.lift(self._main.master) #lifts interfaces above main interface
        
        
    def restart_program(self):
        '''
        Creates popup message to close SimulationUserInputsInterface and enabled start button in MainInterface for user
        to restart simulation inputs.
        '''
        if messagebox.askyesno("Exit", "Are you sure you want to restart the simulation?", parent=self._master):
            self._master.destroy()
            self._main.start_button['state'] = tk.NORMAL
    
    def exit_program(self):
        '''
        Creates a popup message to close entire program.
        '''
        if messagebox.askyesno("Exit", "Are you sure you want to quit the program?", parent=self._main.master):
            self._main.master.destroy()
            
    def on_exit(self):
        '''
        Creates a popup message if "x" is clicked on master.
        '''
        if messagebox.askyesno("Exit", "Are you sure you want to quit the simulation?", parent=self._master):
            self._master.destroy()
            self._main.start_button['state'] = tk.NORMAL

        
            
    
    
    
    
    
    
    